"""合成样本生成 —— 供 cleaner/端到端测试使用。

用 openpyxl 生成 xlsx fixture，reportlab 生成 PDF fixture（dev 依赖）。
样本同时写入 tests/fixtures/（测试用）与 knowledge_base/raw/（CLI 端到端用）。
"""

from __future__ import annotations

from pathlib import Path

import pytest

FIXTURES = Path(__file__).parent / "fixtures"
RAW = Path(__file__).resolve().parents[1] / "l1_kb" / "knowledge_base" / "raw"


def _ensure_dirs() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    for sub in ("data_table", "data_product", "process"):
        (RAW / sub).mkdir(parents=True, exist_ok=True)


def make_order_xlsx(path: Path) -> Path:
    """2-sheet xlsx：订单主表 + 订单明细（PRD §6.4 示例）。"""
    from openpyxl import Workbook

    _ensure_dirs()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "订单"
    ws1.append(["order_id", "customer", "amount", "status"])
    ws1.append(["O1001", "张三", 199.0, "已支付"])
    ws1.append(["O1002", "李四", 88.5, "待发货"])
    ws2 = wb.create_sheet("订单明细")
    ws2.append(["item_id", "order_id", "product", "qty"])
    ws2.append(["I1", "O1001", "笔记本", 2])
    ws2.append(["I2", "O1001", "鼠标", 1])
    ws2.append(["I3", "O1002", "键盘", 1])
    wb.save(path)
    return path


def make_wide_xlsx(path: Path) -> Path:
    """宽表 >20 列，触发 F4 分组/截断。"""
    from openpyxl import Workbook

    _ensure_dirs()
    wb = Workbook()
    ws = wb.active
    ws.title = "宽表"
    # 三组前缀：order_ / amount_ / time_
    headers = (
        [f"order_{c}" for c in ["id", "name", "status"]]
        + [f"amount_{c}" for c in ["total", "tax", "discount"]]
        + [f"time_{c}" for c in ["create", "pay", "ship"]]
        + [f"extra_{i}" for i in range(15)]  # 凑到 >20 列
    )
    ws.append(headers)
    ws.append(["v"] * len(headers))
    wb.save(path)
    return path


def make_pdf(path: Path) -> Path:
    """用 reportlab 生成含标题 + 表格的 PDF。"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

    _ensure_dirs()
    doc = SimpleDocTemplate(str(path), pagesize=A4)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Product Introduction", styles["Title"]),
        Paragraph("This document describes the product.", styles["BodyText"]),
        Paragraph("Field Specification", styles["Heading2"]),
        Table(
            [["field", "type", "desc"], ["order_id", "string", "订单号"], ["amount", "number", "金额"]],
            style=TableStyle(),
        ),
    ]
    doc.build(story)
    return path


@pytest.fixture(scope="session")
def order_xlsx() -> Path:
    p = FIXTURES / "order_detail.xlsx"
    if not p.exists():
        make_order_xlsx(p)
    return p


@pytest.fixture(scope="session")
def wide_xlsx() -> Path:
    p = FIXTURES / "wide_table.xlsx"
    if not p.exists():
        make_wide_xlsx(p)
    return p


@pytest.fixture(scope="session")
def sample_pdf() -> Path:
    p = FIXTURES / "product_intro.pdf"
    if not p.exists():
        make_pdf(p)
    return p
